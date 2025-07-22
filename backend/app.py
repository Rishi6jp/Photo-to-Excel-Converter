from fastapi import FastAPI, File, UploadFile
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
import os
import uuid
import time

app = FastAPI()

# CORS: Allow requests from any origin (for dev; restrict in prod)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/upload-image/")
async def upload_file(file: UploadFile = File(...)):
    total_start = time.time()

    contents = await file.read()
    read_time = time.time()
    print("✅ [TIME] File read in:", read_time - total_start, "seconds")

    try:
        img = Image.open(io.BytesIO(contents))
        img.load()
        img = img.convert("RGB")
    except Exception as e:
        return {"error": f"Failed to open image: {str(e)}"}

    load_time = time.time()
    print("✅ [TIME] Image loaded and converted in:", load_time - read_time, "seconds")

    # Resize image with nearest-neighbor (preserve pixel art look)
    scale_factor = 0.35 # ⬅️ Slightly reduced for performance
    img = img.resize(
        (max(1, int(img.width * scale_factor)), max(1, int(img.height * scale_factor))),
        Image.NEAREST
    )
    img = img.quantize(colors=32).convert("RGB")


    resize_time = time.time()
    print("✅ [TIME] Image resized in:", resize_time - load_time, "seconds")

    pixels = img.load()
    width, height = img.size

    wb = Workbook()
    ws = wb.active

    excel_start = time.time()
    fill_cache = {}

    for y in range(height):
        for x in range(width):
            r, g, b = pixels[x, y]
            hex_color = f"{r:02X}{g:02X}{b:02X}"

            if hex_color not in fill_cache:
                fill_cache[hex_color] = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid"
                )

            cell = ws.cell(row=y + 1, column=x + 1)
            cell.fill = fill_cache[hex_color]

    pixel_time = time.time()
    print("✅ [TIME] Excel pixel writing done in:", pixel_time - excel_start, "seconds")

    # Set row height and column width for pixel shape
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 1.2
    for row in range(1, height + 1):
        ws.row_dimensions[row].height = 7

    format_time = time.time()
    print("✅ [TIME] Excel formatting done in:", format_time - pixel_time, "seconds")

    # Save Excel to temporary directory
    output_filename = f"pixel_art_{uuid.uuid4().hex}.xlsx"
    output_path = os.path.join("/tmp", output_filename)
    wb.save(output_path)

    save_time = time.time()
    print("✅ [TIME] Excel saved in:", save_time - format_time, "seconds")
    print("✅ [TOTAL TIME] Full processing done in:", save_time - total_start, "seconds")

    return FileResponse(
        path=output_path,
        filename="pixel_art.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
