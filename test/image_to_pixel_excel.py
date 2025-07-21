from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# === Configurations ===
INPUT_IMAGE = "input.jpg"
OUTPUT_EXCEL = "pixel_art.xlsx"
RESIZE_TO = (250, 250)  # width x height - smaller means more "pixel" style

# === Step 1: Load and Resize Image ===
img = Image.open(INPUT_IMAGE).convert('RGB')
img = img.resize(RESIZE_TO, Image.NEAREST)  # pixelate
pixels = img.load()
width, height = img.size

# === Step 2: Create Excel Workbook ===
wb = Workbook()
ws = wb.active

# === Step 3: Fill Excel Cells with Pixel Colors ===
for y in range(height):
    for x in range(width):
        r, g, b = pixels[x, y]
        hex_color = f'{r:02X}{g:02X}{b:02X}'
        cell = ws.cell(row=y+1, column=x+1)
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

# Adjust column widths and row heights to make cells square-ish
for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = 2

for row in range(1, height + 1):
    ws.row_dimensions[row].height = 15

# === Step 4: Save Workbook ===
wb.save(OUTPUT_EXCEL)
print(f"✅ Excel pixel art saved to {OUTPUT_EXCEL}")
